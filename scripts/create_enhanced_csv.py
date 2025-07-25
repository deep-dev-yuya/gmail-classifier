#!/usr/bin/env python3
"""
retraining_candidates_sheet.csvに正解ラベル列を追加し、
Excelでドロップダウン選択できるように改良するスクリプト
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_enhanced_retraining_csv():
    """
    既存のCSVファイルに正解ラベル列を追加し、Excelファイルを作成
    """
    print("=== 改良版retraining_candidates_sheet作成開始 ===\n")
    
    # 既存CSVファイルのパス
    original_csv_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_sheet.csv'
    
    # 新しいファイルのパス
    enhanced_csv_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.csv'
    enhanced_xlsx_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.xlsx'
    
    if not os.path.exists(original_csv_path):
        print(f"Error: Original file not found: {original_csv_path}")
        return
    
    # 既存データ読み込み
    df = pd.read_csv(original_csv_path)
    print(f"Original data loaded: {len(df)} records")
    print(f"Columns: {list(df.columns)}")
    
    # 正解ラベル列を追加（空で初期化）
    if 'groundTruth' not in df.columns:
        df['groundTruth'] = ''
        print("Added 'groundTruth' column")
    
    # コメント列を追加（正解理由を記録用）
    if 'correctionReason' not in df.columns:
        df['correctionReason'] = ''
        print("Added 'correctionReason' column")
    
    # 修正日時列を追加
    if 'correctedAt' not in df.columns:
        df['correctedAt'] = ''
        print("Added 'correctedAt' column")
    
    # 修正者列を追加
    if 'correctedBy' not in df.columns:
        df['correctedBy'] = ''
        print("Added 'correctedBy' column")
    
    # 列の順序を整理
    column_order = [
        'timestamp', 'messageId', 'subject', 
        'predictedClass', 'confidence', 'reason',
        'groundTruth', 'correctionReason', 'correctedAt', 'correctedBy'
    ]
    
    # 存在する列のみを使用
    available_columns = [col for col in column_order if col in df.columns]
    df_ordered = df[available_columns]
    
    # CSV形式で保存
    df_ordered.to_csv(enhanced_csv_path, index=False, encoding='utf-8')
    print(f"Enhanced CSV saved: {enhanced_csv_path}")
    
    # Excel形式で保存（ドロップダウン付き）
    create_excel_with_dropdown(df_ordered, enhanced_xlsx_path)
    
    return df_ordered

def create_excel_with_dropdown(df, xlsx_path):
    """
    ドロップダウン選択機能付きExcelファイルを作成
    """
    print("Creating Excel file with dropdown validation...")
    
    # Workbook作成
    wb = Workbook()
    ws = wb.active
    ws.title = "Gmail分類正解ラベル付け"
    
    # データフレームをワークシートに書き込み
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # スタイル設定
    setup_excel_styles(ws)
    
    # ドロップダウン設定
    setup_dropdown_validation(ws, df)
    
    # 列幅自動調整
    adjust_column_widths(ws)
    
    # 保存
    wb.save(xlsx_path)
    print(f"Enhanced Excel file saved: {xlsx_path}")

def setup_excel_styles(ws):
    """
    Excelファイルのスタイルを設定
    """
    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # データ行の交互背景色
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row_num in range(2, ws.max_row + 1):
        if row_num % 2 == 0:
            for col_num in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col_num).fill = light_fill
    
    # フォント設定
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="游ゴシック", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def setup_dropdown_validation(ws, df):
    """
    正解ラベル列にドロップダウン検証を設定
    """
    # 分類ラベルの定義
    classification_labels = [
        "支払い関係",
        "重要", 
        "プロモーション",
        "仕事・学習"
    ]
    
    # groundTruth列を特定
    groundtruth_col = None
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value == 'groundTruth':
            groundtruth_col = col_num
            break
    
    if groundtruth_col is None:
        print("Warning: groundTruth column not found")
        return
    
    # ドロップダウン検証を作成
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(classification_labels)}"',
        allow_blank=True
    )
    dv.error = '無効な値です'
    dv.errorTitle = '入力エラー'
    dv.prompt = '正解ラベルを選択してください'
    dv.promptTitle = 'ラベル選択'
    
    # groundTruth列の全行に適用
    col_letter = ws.cell(row=1, column=groundtruth_col).column_letter
    range_string = f"{col_letter}2:{col_letter}{ws.max_row + 100}"  # 将来の行も考慮
    dv.add(range_string)
    
    ws.add_data_validation(dv)
    
    # 正解ラベル列のヘッダーを強調
    header_cell = ws.cell(row=1, column=groundtruth_col)
    header_cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_cell.font = Font(color="FFFFFF", bold=True)
    
    print(f"Dropdown validation added to column {col_letter} (groundTruth)")

def adjust_column_widths(ws):
    """
    列幅を内容に合わせて自動調整
    """
    column_widths = {
        'timestamp': 20,
        'messageId': 15,
        'subject': 50,
        'predictedClass': 15,
        'confidence': 12,
        'reason': 15,
        'groundTruth': 15,
        'correctionReason': 30,
        'correctedAt': 20,
        'correctedBy': 15
    }
    
    for col_num, cell in enumerate(ws[1], 1):
        col_name = cell.value
        if col_name in column_widths:
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = column_widths[col_name]

def create_instruction_sheet():
    """
    使用方法の説明シートを作成
    """
    instructions_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/正解ラベル付け_使用方法.md'
    
    instructions = """# Gmail分類正解ラベル付け - 使用方法

## 📋 概要
`retraining_candidates_enhanced.xlsx`ファイルを使用して、Gmail分類の正解ラベルを効率的に付与するためのガイドです。

## 🚀 使用手順

### 1. ファイルを開く
- `n8n/retraining_candidates_enhanced.xlsx` をExcelで開く
- 「Gmail分類正解ラベル付け」シートを確認

### 2. 正解ラベルを付与
- **groundTruth列**（オレンジ色のヘッダー）をクリック
- ドロップダウンから正しいラベルを選択：
  - `支払い関係`
  - `重要`
  - `プロモーション` 
  - `仕事・学習`

### 3. 修正理由を記録（任意）
- **correctionReason列**に修正理由を入力
- 例：「デビットカード通知のため支払い関係が正解」

### 4. 修正情報を記録（任意）
- **correctedAt**: 修正日時（例：2025-07-24）
- **correctedBy**: 修正者名（例：ユーザー名）

## 📊 各列の説明

| 列名 | 説明 |
|------|------|
| timestamp | メール処理日時 |
| messageId | メール識別ID |
| subject | メール件名 |
| predictedClass | AIの予測ラベル |
| confidence | 予測信頼度 |
| reason | AI判定理由 |
| **groundTruth** | **正解ラベル（要入力）** |
| correctionReason | 修正理由（任意） |
| correctedAt | 修正日時（任意） |
| correctedBy | 修正者（任意） |

## 🎯 ラベル付けのガイドライン

### 支払い関係
- PayPay、デビットカード、クレジットカード利用通知
- 料金請求、引き落とし通知
- 銀行振込、口座振替関連

### 重要
- システム障害、メンテナンス通知
- セキュリティアラート
- 契約更新、期限切れ通知
- 法定点検、重要手続き

### プロモーション
- セール、キャンペーン情報
- 割引、特価情報
- マーケティングメール
- 商品・サービス宣伝

### 仕事・学習
- 求人情報、転職関連
- 技術情報、プログラミング
- 学習コンテンツ
- 業務関連通知

## 💡 効率化のコツ

1. **一括処理**: 似たようなメールは連続して処理
2. **パターン認識**: 件名のパターンから素早く判定
3. **理由記録**: 判断に迷った場合は理由を記録
4. **定期保存**: 作業中は定期的に保存

## 🔄 データ活用フロー

1. Excel で正解ラベル付与
2. CSVとして保存・エクスポート
3. `analyze_groundtruth_data.py` で機械学習
4. 新モデルをFlask APIに統合
5. 分類精度の向上を確認

## ⚠️ 注意事項

- **groundTruth列**は必須入力項目
- ドロップダウン以外の値は入力しない
- ファイルの破損を防ぐため定期的にバックアップ
- 大量データ処理時はExcelの動作が重くなる場合あり

## 📞 サポート

分類判断に迷った場合：
1. 既存の正解データを参考にする
2. メール内容の文脈を重視する
3. 不明な場合は「重要」として保守的に分類

---
*作成日: 2025-07-24*
*ファイル場所: `/Users/user/Projects/dev-projects/gmail-classifier/n8n/`*
"""
    
    with open(instructions_path, 'w', encoding='utf-8') as f:
        f.write(instructions)
    
    print(f"Instructions created: {instructions_path}")

if __name__ == "__main__":
    # 改良版CSVとExcelファイルを作成
    df_enhanced = create_enhanced_retraining_csv()
    
    # 使用方法説明書を作成
    create_instruction_sheet()
    
    print("\n=== 作成完了 ===")
    print("📁 作成されたファイル:")
    print("1. retraining_candidates_enhanced.csv - 改良版CSV")
    print("2. retraining_candidates_enhanced.xlsx - ドロップダウン付きExcel")
    print("3. 正解ラベル付け_使用方法.md - 使用方法ガイド")
    print("\n🚀 次のステップ:")
    print("1. retraining_candidates_enhanced.xlsx をExcelで開く")
    print("2. groundTruth列（オレンジ色）でドロップダウンから正解ラベルを選択")
    print("3. 修正理由をcorrectionReason列に記録（任意）")
    print("4. 完了後、analyze_groundtruth_data.py で新モデル学習")