#!/usr/bin/env python3
"""
CSV（n8n自動書き込み）とExcel（ドロップダウン付き）を同期するスクリプト
"""

import pandas as pd
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font

def sync_csv_to_excel():
    """
    CSVファイルの新しいデータをExcelファイルに同期
    """
    print("=== CSV → Excel 同期開始 ===\n")
    
    csv_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.csv'
    excel_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.xlsx'
    
    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        return False
    
    # CSVファイルを読み込み
    df_csv = pd.read_csv(csv_path)
    print(f"CSV records: {len(df_csv)}")
    
    # Excelファイルの存在確認
    if os.path.exists(excel_path):
        # 既存Excelファイルを読み込み（正解ラベル付き）
        df_excel = pd.read_excel(excel_path)
        print(f"Excel records: {len(df_excel)}")
        
        # 既存のラベル付きデータを保持
        labeled_data = {}
        for _, row in df_excel.iterrows():
            if pd.notna(row.get('groundTruth')) and row.get('groundTruth') != '':
                labeled_data[row['messageId']] = {
                    'groundTruth': row['groundTruth'],
                    'correctionReason': row.get('correctionReason', ''),
                    'correctedAt': row.get('correctedAt', ''),
                    'correctedBy': row.get('correctedBy', '')
                }
        
        print(f"Existing labeled records: {len(labeled_data)}")
    else:
        labeled_data = {}
        print("No existing Excel file found")
    
    # CSVデータに既存ラベルをマージ
    for idx, row in df_csv.iterrows():
        message_id = row['messageId']
        if message_id in labeled_data:
            df_csv.at[idx, 'groundTruth'] = labeled_data[message_id]['groundTruth']
            df_csv.at[idx, 'correctionReason'] = labeled_data[message_id]['correctionReason']
            df_csv.at[idx, 'correctedAt'] = labeled_data[message_id]['correctedAt']
            df_csv.at[idx, 'correctedBy'] = labeled_data[message_id]['correctedBy']
    
    # 新しいレコード数を計算
    new_records = len(df_csv) - len(labeled_data)
    print(f"New records to add: {new_records}")
    
    # 更新されたExcelファイルを作成
    create_enhanced_excel(df_csv, excel_path)
    
    print(f"✅ Sync completed: {len(df_csv)} total records")
    return True

def sync_excel_to_csv():
    """
    Excelファイルの正解ラベルをCSVファイルに反映
    """
    print("=== Excel → CSV 同期開始 ===\n")
    
    csv_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.csv'
    excel_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/retraining_candidates_enhanced.xlsx'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        return False
    
    if not os.path.exists(csv_path):
        print(f"Error: CSV file not found: {csv_path}")
        return False
    
    # ファイルを読み込み
    df_excel = pd.read_excel(excel_path)
    df_csv = pd.read_csv(csv_path)
    
    print(f"Excel records: {len(df_excel)}")
    print(f"CSV records: {len(df_csv)}")
    
    # Excelから正解ラベル情報を抽出
    labeled_count = 0
    for _, excel_row in df_excel.iterrows():
        message_id = excel_row['messageId']
        
        # CSVの対応行を検索
        csv_idx = df_csv[df_csv['messageId'] == message_id].index
        
        if len(csv_idx) > 0:
            idx = csv_idx[0]
            
            # 正解ラベルが入力されている場合のみ更新
            if pd.notna(excel_row.get('groundTruth')) and excel_row.get('groundTruth') != '':
                df_csv.at[idx, 'groundTruth'] = excel_row['groundTruth']
                df_csv.at[idx, 'correctionReason'] = excel_row.get('correctionReason', '')
                df_csv.at[idx, 'correctedAt'] = excel_row.get('correctedAt', '')
                df_csv.at[idx, 'correctedBy'] = excel_row.get('correctedBy', '')
                labeled_count += 1
    
    # CSVファイルを更新
    df_csv.to_csv(csv_path, index=False, encoding='utf-8')
    
    print(f"✅ Sync completed: {labeled_count} labels synced to CSV")
    return True

def create_enhanced_excel(df, excel_path):
    """
    データ検証（ドロップダウン）付きExcelファイルを作成
    """
    print("Creating enhanced Excel with dropdowns...")
    
    # バックアップ作成
    if os.path.exists(excel_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = excel_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
        shutil.copy2(excel_path, backup_path)
        print(f"Backup created: {os.path.basename(backup_path)}")
    
    # Excelファイルに書き込み
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Gmail分類正解ラベル付け', index=False)
        
        # ワークシートを取得
        worksheet = writer.sheets['Gmail分類正解ラベル付け']
        
        # ヘッダースタイル設定
        setup_header_styles(worksheet)
        
        # ドロップダウン設定
        setup_dropdown_validation(worksheet, len(df))
        
        # 列幅調整
        adjust_column_widths(worksheet)

def setup_header_styles(ws):
    """ヘッダー行のスタイルを設定"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # groundTruth列を特別に強調
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value == 'groundTruth':
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            break

def setup_dropdown_validation(ws, data_rows):
    """groundTruth列にドロップダウン検証を設定"""
    classification_labels = [
        "支払い関係",
        "重要", 
        "プロモーション",
        "仕事・学習"
    ]
    
    # groundTruth列を特定
    groundtruth_col = None
    for col_num, cell in enumerate(ws[1], 1):
        if cell.value == 'groundTruth':
            groundtruth_col = col_num
            break
    
    if groundtruth_col is None:
        return
    
    # ドロップダウン検証を作成
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(classification_labels)}"',
        allow_blank=True
    )
    dv.error = '無効な値です。正しいラベルを選択してください。'
    dv.errorTitle = '入力エラー'
    dv.prompt = 'ドロップダウンから正解ラベルを選択してください'
    dv.promptTitle = 'ラベル選択'
    
    # 全データ行に適用
    col_letter = ws.cell(row=1, column=groundtruth_col).column_letter
    range_string = f"{col_letter}2:{col_letter}{data_rows + 10}"  # 余裕を持たせる
    dv.add(range_string)
    
    ws.add_data_validation(dv)
    print(f"Dropdown validation added to column {col_letter}")

def adjust_column_widths(ws):
    """列幅を調整"""
    column_widths = {
        'A': 20,  # timestamp
        'B': 15,  # messageId
        'C': 50,  # subject
        'D': 15,  # predictedClass
        'E': 12,  # confidence
        'F': 15,  # reason
        'G': 15,  # groundTruth
        'H': 30,  # correctionReason
        'I': 20,  # correctedAt
        'J': 15   # correctedBy
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def create_sync_instructions():
    """同期操作の手順書を作成"""
    instructions = """# CSV-Excel同期操作手順書

## 🔄 同期の種類

### **1. CSV → Excel 同期（新データ追加）**

n8nでCSVに追加されたデータをExcelに反映する際に使用：

```bash
python3 scripts/sync_csv_excel.py --csv-to-excel
```

**実行タイミング**:
- n8nで新しいメールが処理された後
- ラベル付け作業を開始する前

**処理内容**:
- CSVの新しいレコードをExcelに追加
- 既存のラベル付きデータを保持
- ドロップダウン機能を再設定

### **2. Excel → CSV 同期（ラベル反映）**

Excelでラベル付けした結果をCSVに反映する際に使用：

```bash
python3 scripts/sync_csv_excel.py --excel-to-csv
```

**実行タイミング**:
- ラベル付け作業完了後
- モデル更新を実行する前

**処理内容**:
- Excelの正解ラベルをCSVに転写
- モデル学習用データを準備

---

## 🚀 推奨運用フロー

### **日次運用**
1. n8nがCSVに自動書き込み
2. 手動で CSV → Excel 同期実行
3. Excelでラベル付け作業
4. Excel → CSV 同期実行

### **週次運用**
1. モデル更新実行
2. API反映
3. 精度確認

---

## ⚠️ 注意事項

- 同期前に必ずバックアップが作成されます
- Excelファイルでの作業中は同期を実行しないでください
- エラーが発生した場合はバックアップから復元可能です

---

*作成日: 2025-07-24*
"""
    
    instructions_path = '/Users/user/Projects/dev-projects/gmail-classifier/n8n/CSV-Excel同期手順書.md'
    with open(instructions_path, 'w', encoding='utf-8') as f:
        f.write(instructions)
    
    return instructions_path

def main():
    """メイン実行関数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='CSV-Excel同期スクリプト')
    parser.add_argument('--csv-to-excel', action='store_true', help='CSVからExcelに同期')
    parser.add_argument('--excel-to-csv', action='store_true', help='ExcelからCSVに同期')
    parser.add_argument('--both', action='store_true', help='双方向同期（CSV→Excel→CSV）')
    
    args = parser.parse_args()
    
    if args.both:
        print("双方向同期を実行します...")
        sync_csv_to_excel()
        print("\n" + "="*50 + "\n")
        sync_excel_to_csv()
    elif args.csv_to_excel:
        sync_csv_to_excel()
    elif args.excel_to_csv:
        sync_excel_to_csv()
    else:
        # 引数なしの場合は双方向同期
        print("引数が指定されていません。双方向同期を実行します...")
        sync_csv_to_excel()
        print("\n" + "="*50 + "\n")
        
        response = input("Excelでラベル付け作業は完了していますか？ (y/N): ").strip().lower()
        if response in ['y', 'yes']:
            sync_excel_to_csv()
        else:
            print("ラベル付け作業完了後に以下を実行してください：")
            print("python3 scripts/sync_csv_excel.py --excel-to-csv")
    
    # 手順書作成
    instructions_path = create_sync_instructions()
    print(f"\n📋 同期手順書を作成: {os.path.basename(instructions_path)}")

if __name__ == "__main__":
    main()